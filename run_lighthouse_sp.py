import subprocess
import os
import json
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# 計測するURLリスト
urls = [
    "https://furusato.jreast.co.jp/furusato",  # URL A
    "https://furusato.jreast.co.jp/furusato/ranking"   # URL B
]

# Lighthouseの計測とExcelへの保存を行う関数
def run_lighthouse_for_url(url, output_dir):
    os.makedirs(output_dir, exist_ok=True)  # 出力ディレクトリを作成

    # Lighthouseのプリセット設定
    for i in range(1, 6):  # 5回計測
        output_path_json = os.path.abspath(os.path.join(output_dir, f"report_{i}.json"))
        log_path = os.path.abspath(os.path.join(output_dir, f"lighthouse_verbose_run_{i}.log"))

        print(f"Running Lighthouse for run {i} on URL: {url}")
        print(f"Output JSON path: {output_path_json}")
        print(f"Log path: {log_path}")

        # Lighthouseを実行するコマンド
        command_json = [
            "npx", "lighthouse", url,
            "--preset=perf",
            "--output", "json", "html",
            "--output-path", output_path_json,
            "--chrome-flags=\"--headless --no-sandbox\"",
            "--max-wait-for-load=60000",
            "--verbose",
            "--emulated-form-factor=mobile",
            "--throttling-method=simulate",
            "--throttling.cpuSlowdownMultiplier=2",
            "--throttling.throughputKbps=6000",
            "--throttling.uploadThroughputKbps=750",
            "--throttling.latency=100"
        ]
        with open(log_path, "w") as log_file:
            result_json = subprocess.run(command_json, stdout=log_file, stderr=subprocess.STDOUT)

        if result_json.returncode == 0:
            print(f"Lighthouse JSON run {i} completed successfully.")
        else:
            print(f"Error in Lighthouse JSON run {i}. Check {log_path} for details.")
            break

    # 指標の抽出とExcel保存
    save_metrics_to_excel(output_dir, url)

def save_metrics_to_excel(output_dir, url):
    metrics_to_extract = [
        "FCP(ms)",
        "LCP(ms)",
        "CLS",
        "Speed Index(ms)",
        "TBT(ms)",
        "TTI(ms)",
        "TTFB(ms)",
        "Performance"
    ]

    results = []
    for i in range(1, 6):  # JSONファイルが5つあると仮定
        json_path = os.path.join(output_dir, f"report_{i}.json")
        if os.path.exists(json_path):
            with open(json_path, 'r') as f:
                data = json.load(f)
                try:
                    audits = data.get("audits", {})
                    performance_score = data.get("categories", {}).get("performance", {}).get("score", "N/A")
                    result = {
                        "run": i,
                        "Performance": round(performance_score * 100, 2) if isinstance(performance_score, (int, float)) else "N/A",
                        "TTFB(ms)": audits.get("server-response-time", {}).get("numericValue", "N/A"),
                        "FCP(ms)": audits.get("first-contentful-paint", {}).get("numericValue", "N/A"),
                        "LCP(ms)": audits.get("largest-contentful-paint", {}).get("numericValue", "N/A"),
                        "Speed Index(ms)": audits.get("speed-index", {}).get("numericValue", "N/A"),
                        "TBT(ms)": audits.get("total-blocking-time", {}).get("numericValue", audits.get("total-blocking-time", {}).get("errorMessage", "N/A")),
                        "TTI(ms)": audits.get("interactive", {}).get("numericValue", audits.get("interactive", {}).get("errorMessage", "N/A")),
                        "CLS": audits.get("cumulative-layout-shift", {}).get("numericValue", "N/A")
                    }
                    results.append(result)
                except KeyError as e:
                    print(f"Error extracting metrics from {json_path}: {e}")
        else:
            print(f"JSON file {json_path} not found!")

    # DataFrameを作成
    if results:
        df = pd.DataFrame(results)

        # 平均値を計算
        averages = pd.DataFrame([df.mean(numeric_only=True)]).assign(run="average")
        df = pd.concat([df, averages], ignore_index=True)

        # 実行完了日時を取得
        completed_datetime = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # Excelファイルの書き出し
        excel_path = os.path.join(output_dir, f"{output_dir}.xlsx")
        with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
            # データを出力
            df.to_excel(writer, index=False, startrow=2, sheet_name="Metrics")  # 2行目から書き始める
            workbook = writer.book
            worksheet = writer.sheets["Metrics"]

            # URLと実行日時を追加
            worksheet["A1"] = f"URL: {url}"
            worksheet["A2"] = f"Completed at: {completed_datetime}"

            # 列幅の調整
            for col_num in range(2, 9):  # B〜H
                col_letter = get_column_letter(col_num)
                worksheet.column_dimensions[col_letter].width = 15

            # 枠線を設定
            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )
            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.row > 2:  # データ部分に枠線を設定
                        cell.border = thin_border

            # 左揃えの設定（A列と1行目）
            for cell in worksheet["A"]:
                cell.alignment = Alignment(horizontal="left")

        print(f"計測結果のExcelファイルが次の場所に保存されました💫: {excel_path}")
    else:
        print("No metrics data found. Excel file not created.")

# メイン処理
if __name__ == "__main__":
    for url in urls:
        # URLごとにディレクトリを作成
        output_dir = url.split("//")[1].replace("/", "_")  # URLを安全なディレクトリ名に変換
        run_lighthouse_for_url(url, output_dir)
